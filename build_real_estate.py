"""
Constrói o real_estate_data.json a partir da planilha upload.xlsx
Agrupa unidades por endereço-base (rua/avenida) criando "prédios" automáticos.
"""
import json, re, os
import openpyxl

XLSX_PATH = os.path.join(os.path.dirname(__file__), 'upload.xlsx')
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), 'real_estate_data.json')

wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws = wb.active

# ── 1. Ler cabeçalhos ────────────────────────────────────────────────────────
row1 = [cell.value for cell in ws[1]]  # nomes dos meses
row2 = [cell.value for cell in ws[2]]  # descrições das colunas

# Detectar colunas de meses na row1
PT_MONTHS = {
    'janeiro':'01','jan':'01','fevereiro':'02','fev':'02',
    'março':'03','marco':'03','mar':'03','abril':'04','abr':'04',
    'maio':'05','mai':'05','junho':'06','jun':'06',
    'julho':'07','jul':'07','agosto':'08','ago':'08',
    'setembro':'09','set':'09','outubro':'10','out':'10',
    'novembro':'11','nov':'11','dezembro':'12','dez':'12'
}

def parse_month_label(label):
    if not label: return None
    lc = str(label).lower()
    yr_match = re.search(r'20\d{2}', lc)
    if not yr_match: return None
    yr = yr_match.group()
    mo = None
    for k, v in PT_MONTHS.items():
        if k in lc:
            mo = v
            break
    if not mo:
        mm = re.search(r'(\d{2})/', lc)
        if mm: mo = mm.group(1)
    if mo and yr: return f"{mo}/{yr}"
    return None

month_cols = {}  # col_index -> "MM/YYYY"
for i, cell_val in enumerate(row1):
    parsed = parse_month_label(cell_val)
    if parsed:
        month_cols[i] = parsed

print(f"Meses detectados: {month_cols}")

# Detectar colunas descritivas na row2
col_address = col_tenant = col_status = col_open_rent = col_open_sale = -1
for i, cell_val in enumerate(row2):
    h = str(cell_val or '').lower()
    if 'endere' in h or 'imovel' in h or 'imóvel' in h: col_address = i
    if 'locata' in h or 'locater' in h or 'nome do loc' in h: col_tenant = i
    if 'locado' in h or h == 'status' or 'situa' in h: col_status = i
    if 'aberto para loca' in h: col_open_rent = i
    if 'aberto para vend' in h: col_open_sale = i

print(f"Colunas: Endereço={col_address}, Inquilino={col_tenant}, Status={col_status}, AbertoLoc={col_open_rent}, AbertoVend={col_open_sale}")

# ── 2. Ler todas as linhas de dados ──────────────────────────────────────────
units_raw = []
for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
    cells = list(row)
    address = str(cells[col_address] or '').strip() if col_address >= 0 and col_address < len(cells) else ''
    tenant = str(cells[col_tenant] or '').strip() if col_tenant >= 0 and col_tenant < len(cells) else ''
    status_raw = str(cells[col_status] or '').strip() if col_status >= 0 and col_status < len(cells) else ''
    
    if not address and not tenant:
        continue
    if status_raw.lower() == 'total':
        continue

    # Parse status
    s = status_raw.lower()
    if 'alugado' in s: mapped_status = 'alugado'
    elif 'devendo' in s: mapped_status = 'alugado'  # devendo = still rented
    elif 'disponivel' in s or 'disponível' in s: mapped_status = 'disponivel'
    elif 'preparando' in s: mapped_status = 'disponivel'
    elif 'vender' in s: mapped_status = 'disponivel'  # à venda
    elif 'vendido' in s: mapped_status = 'vendido'
    else: mapped_status = 'disponivel'

    # Parse monthly values
    monthly = {}
    for ci, mmyyyy in month_cols.items():
        if ci < len(cells):
            val = cells[ci]
            if isinstance(val, (int, float)):
                monthly[mmyyyy] = round(val, 2)
            elif val:
                try:
                    clean = str(val).replace('R$','').replace(' ','').replace('.','').replace(',','.')
                    monthly[mmyyyy] = round(float(clean), 2)
                except: pass

    # Parse open rent / sale values
    open_rent = 0
    open_sale = 0
    if col_open_rent >= 0 and col_open_rent < len(cells) and cells[col_open_rent]:
        try: open_rent = float(cells[col_open_rent]) if isinstance(cells[col_open_rent], (int,float)) else 0
        except: pass
    if col_open_sale >= 0 and col_open_sale < len(cells) and cells[col_open_sale]:
        try: open_sale = float(cells[col_open_sale]) if isinstance(cells[col_open_sale], (int,float)) else 0
        except: pass

    # Compute rent value: use the most recent month with a positive value
    rent_value = 0
    if mapped_status == 'alugado':
        for mmyyyy in sorted(monthly.keys(), reverse=True):
            if monthly[mmyyyy] > 0:
                rent_value = monthly[mmyyyy]
                break
    
    # Sale value for "vender" status
    sale_value = open_sale if open_sale > 0 else 0

    # Parse notes from last column (J) if present
    notes_col = 9  # column J
    notes = ''
    if notes_col < len(cells) and cells[notes_col]:
        notes = str(cells[notes_col]).strip()

    units_raw.append({
        'address': address,
        'tenant': tenant,
        'status': mapped_status,
        'status_label': status_raw,
        'rent_value': rent_value,
        'sale_value': sale_value,
        'open_rent': open_rent,
        'open_sale': open_sale,
        'monthly': monthly,
        'notes': notes
    })

print(f"\nTotal de unidades lidas: {len(units_raw)}")

# ── 3. Agrupar por "prédio" baseado no endereço ─────────────────────────────
def get_building_key(address):
    """Extrai o nome do 'prédio' a partir do endereço."""
    addr = address.strip()
    if not addr:
        return 'Sem Endereço'
    
    # Padrões especiais
    if addr.upper().startswith('EPG'):
        return 'EPG - Eng. Pedro Garcin (Kitnet)'
    if 'odzun' in addr.lower():
        return 'Prédio Odzun'
    
    # Extrair a rua base (remover número do imóvel, sala, casa, apt, etc.)
    # Ex: "Rua Alfredo da Costa Figo 41 Campinas" -> "Rua Alfredo da Costa Figo - Campinas"
    # Ex: "Av. João Jose Gomes 140 casa 1" -> "Av. João Jose Gomes 140"
    
    # Remove detalhes após o número principal (casa X, sala X, ap X, etc.)
    clean = re.sub(r'\s+(casa|sala|ap\s|apt\s|loja|galp[aã]o|bl\s|bloco)\s*\d*.*$', '', addr, flags=re.IGNORECASE)
    
    # For specific known patterns
    if 'bittencourt' in addr.lower() or 'Bittencourt' in addr:
        return 'R. Bittencourt 141'
    if 'anchieta' in addr.lower():
        return 'Bertioga - Pass. dos Meros'
    if 'nações unidas' in addr.lower() or 'nacoes unidas' in addr.lower() or 'Na\x87' in addr:
        return 'Av. Nações Unidas 14401 - Pq Cidade'
    if 'vital brasil' in addr.lower():
        return 'Av. Dr. Vital Brasil 305'
    if 'corifeu' in addr.lower():
        return 'Av. Corifeu de Azevedo Marques'
    if 'presidente altino' in addr.lower():
        return 'Av. Presidente Altino 1619'
    if 'romeu ferro' in addr.lower():
        return 'Rua Dr. Romeu Ferro'
    if 'joão batista' in addr.lower() or 'joao batista' in addr.lower() or 'Jo\x87' in addr:
        return 'Rua João Batista Morato do Canto'
    if 'alfredo' in addr.lower() and 'costa figo' in addr.lower():
        return 'Rua Alfredo da Costa Figo - Campinas'
    if 'lauro vannucci' in addr.lower():
        return 'Rua Lauro Vannucci - Campinas'
    if 'eng' in addr.lower() and 'garcin' in addr.lower() and '170' in addr:
        return 'Rua Eng. Pedro Garcin 170'
    if 'eng' in addr.lower() and 'garcin' in addr.lower() and '195' in addr:
        return 'Rua Eng. Pedro Garcin 195'
    if ('jo' in addr.lower() or 'João' in addr) and 'gomes' in addr.lower():
        return 'Av. João Jose Gomes 140'
    
    return clean.strip() or addr

# Group
buildings_map = {}
for u in units_raw:
    bkey = get_building_key(u['address'])
    if bkey not in buildings_map:
        buildings_map[bkey] = []
    buildings_map[bkey].append(u)

print(f"Prédios/agrupamentos: {len(buildings_map)}")
for bkey, units in buildings_map.items():
    print(f"  {bkey}: {len(units)} unidades")

# ── 4. Gerar a estrutura real_estate_data.json ───────────────────────────────
real_estate = []
building_id = 1

for bkey, units in buildings_map.items():
    # Deduzir o endereço do primeiro imóvel do grupo
    first_addr = units[0]['address']
    
    building = {
        'id': building_id,
        'name': bkey,
        'address': first_addr,
        'totalUnits': len(units),
        'units': []
    }
    
    for uid, u in enumerate(units, start=1):
        # Deduzir o label (nome curto da unidade)
        label = u['address']
        # Tentar extrair a parte específica (casa X, sala X, ap X)
        match = re.search(r'(casa\s*\d+|sala\s*[\d/]+|ap\s*\d+|loja\s*\d+|galp[aã]o|Galp[aã]o|bl\s*\d+)', label, re.IGNORECASE)
        if match:
            label = match.group(0).strip()
        elif 'EPG' in u['address']:
            label = u['address']  # EPG 01, EPG 02 etc.
        else:
            # Use a parte após o número da rua
            parts = re.split(r'\d{2,}', label, maxsplit=1)
            if len(parts) > 1 and parts[1].strip():
                label = parts[1].strip()
                # Remove city name if at the end
                label = re.sub(r'\s*(campinas|são paulo|bertioga|sp)$', '', label, flags=re.IGNORECASE).strip()
            if not label or len(label) < 2:
                label = u['address']
        
        # For Odzun, extract ap number
        if 'odzun' in u['address'].lower():
            od_match = re.search(r'(ap\s*\d+[^,]*)', u['address'], re.IGNORECASE)
            if od_match:
                label = od_match.group(1).strip()
            else:
                label = u['address'].replace('Odzun ', '').strip()

        unit_obj = {
            'id': uid,
            'label': label,
            'status': u['status'],
            'rentValue': u['rent_value'] if u['status'] == 'alugado' else 0,
            'saleValue': u['sale_value'] if u['status'] == 'vendido' or u['open_sale'] > 0 else 0,
            'tenantName': u['tenant'] if u['tenant'] else '',
            'notes': u['notes'] if u['notes'] else f"Status original: {u['status_label']}",
            'monthlyReceipts': u['monthly'],
            'rentStartDate': '2026-02-01',  # Earliest month in spreadsheet
        }
        
        # For units à venda, store the asking price
        if u['open_sale'] > 0 and u['status'] != 'vendido':
            unit_obj['saleValue'] = u['open_sale']
            unit_obj['notes'] = f"À venda por R$ {u['open_sale']:,.2f}. {u['notes']}".strip()
        
        # For available units with open rent
        if u['open_rent'] > 0:
            unit_obj['rentValue'] = u['open_rent']
            unit_obj['notes'] = f"Pedir R$ {u['open_rent']:,.2f}/mês. {unit_obj['notes']}".strip()

        building['units'].append(unit_obj)
    
    real_estate.append(building)
    building_id += 1

# ── 5. Salvar ─────────────────────────────────────────────────────────────────
with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
    json.dump(real_estate, f, ensure_ascii=False, indent=2)

# Stats
total_units = sum(len(b['units']) for b in real_estate)
total_alugados = sum(1 for b in real_estate for u in b['units'] if u['status'] == 'alugado')
total_vendidos = sum(1 for b in real_estate for u in b['units'] if u['status'] == 'vendido')
total_disponiveis = sum(1 for b in real_estate for u in b['units'] if u['status'] == 'disponivel')

print(f"\n{'='*60}")
print(f"real_estate_data.json gerado com sucesso!")
print(f"  Prédios:     {len(real_estate)}")
print(f"  Unidades:    {total_units}")
print(f"  Alugados:    {total_alugados}")
print(f"  Vendidos:    {total_vendidos}")
print(f"  Disponíveis: {total_disponiveis}")
print(f"{'='*60}")
